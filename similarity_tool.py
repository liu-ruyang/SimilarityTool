import configparser
import os

from openpyxl.reader.excel import load_workbook
from similarities import BertSimilarity


class Tool():
    # 计算text1和text2的文本相似度
    def calculate(self, m, text1, text2):
        result = m.similarity(text1, text2)
        return f"{float(result)}"

    def process(self, AColumn, BColumn, startRow, endRow, resultColumn, model_name_or_path, device, xslxUrl):
        m = BertSimilarity(model_name_or_path=model_name_or_path, device=device)

        self.wb = load_workbook(xslxUrl)
        self.ws = self.wb["Sheet1"]
        rows = self.ws[startRow:endRow]

        count = 0
        print("开始处理…")
        for row in rows:
            text1 = row[int(AColumn)].value  # A列单元格
            text2 = row[int(BColumn)].value  # B列单元格
            cell1 = row[int(resultColumn)]  # 计算结果所在单元格
            # 计算匹配相似度
            result1 = self.calculate(m, text1, text2)
            cell1.value = result1
            count += 1
            print("第", count, "条数据处理结束")
            self.wb.save(xslxUrl)


if __name__ == '__main__':
    # 创建ConfigParser对象
    config = configparser.ConfigParser()

    # 获取当前脚本文件所在的目录
    script_dir = os.path.dirname(__file__)
    # 构建配置文件的相对路径
    config_path = os.path.join(script_dir, '', 'config.properties')
    # 模型文件夹路径
    model_name_or_path = os.path.join(script_dir, '', 'text2vec-base-chinese')

    print(config_path)

    # 读取properties文件
    config.read(config_path)
    device = config.get("section1", "device")
    xlsxUrl = config.get("section1", "xlsxUrl")
    AColumn = config.get("section1", "AColumn")
    BColumn = config.get("section1", "BColumn")
    startRow = config.get("section1", "startRow")
    endRow = config.get("section1", "endRow")
    resultColumn = config.get("section1", "resultColumn")

    tool = Tool()
    tool.process(AColumn=AColumn, BColumn=BColumn, startRow=startRow, endRow=endRow, resultColumn=resultColumn,
                 model_name_or_path=model_name_or_path, device=device, xslxUrl=xlsxUrl)
    print("处理结束")
